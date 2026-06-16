import argparse
import csv
import math
import os
import statistics
import sys
import xml.etree.ElementTree as ET


BAD_TRUE = {"BAD", "TRUE", "1"}
BAD_FALSE = {"GOOD", "OK", "FALSE", "0"}


def normalize_text(value):
    return str(value).strip() if value is not None else ""


def normalize_bad(value):
    text = normalize_text(value).upper()
    if text in BAD_TRUE:
        return "BAD"
    if text in BAD_FALSE:
        return "GOOD"
    return ""


def parse_float(value):
    text = normalize_text(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def read_csv(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def read_xml_spreadsheet(path):
    ns = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}
    tree = ET.parse(path)
    rows = []
    for row in tree.findall(".//ss:Worksheet/ss:Table/ss:Row", ns):
        values = []
        for cell in row.findall("ss:Cell", ns):
            data = cell.find("ss:Data", ns)
            values.append(data.text if data is not None and data.text is not None else "")
        if values:
            rows.append(values)
    if not rows:
        return []
    headers = rows[0]
    return [dict(zip(headers, row + [""] * (len(headers) - len(row)))) for row in rows[1:]]


def read_xlsx(path):
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read .xlsx input files") from exc

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [normalize_text(v) for v in rows[0]]
    result = []
    for row in rows[1:]:
        result.append({headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))})
    return result


def read_table(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return read_csv(path)
    if ext == ".xls":
        return read_xml_spreadsheet(path)
    if ext == ".xlsx":
        return read_xlsx(path)
    raise ValueError("Unsupported input file type: " + ext)


def binomial_p_value_greater(k, n, p0=0.5):
    if n <= 0:
        return None
    probability = 0.0
    for i in range(k, n + 1):
        probability += math.comb(n, i) * (p0 ** i) * ((1.0 - p0) ** (n - i))
    return min(1.0, probability)


def summarize_binary(rows, pred_col, true_col, normalizer=lambda x: normalize_text(x).upper()):
    labeled = []
    for row in rows:
        truth = normalizer(row.get(true_col, ""))
        pred = normalizer(row.get(pred_col, ""))
        if truth:
            labeled.append((pred, truth))

    n = len(labeled)
    k = sum(1 for pred, truth in labeled if pred == truth)
    return {
        "n": n,
        "k": k,
        "accuracy": (k / n) if n else None,
        "p_value_gt_0_5": binomial_p_value_greater(k, n, 0.5) if n else None,
    }


def summarize_points(rows):
    image_errors = []
    success15 = 0
    success20 = 0
    full_labeled = 0
    point_errors = [[] for _ in range(6)]

    for row in rows:
        mean_error = parse_float(row.get("point_mean_error_px", ""))
        labeled_count_text = normalize_text(row.get("point_labeled_count", ""))
        try:
            labeled_count = int(float(labeled_count_text)) if labeled_count_text else 0
        except ValueError:
            labeled_count = 0

        if mean_error is not None and labeled_count > 0:
            image_errors.append(mean_error)

        if labeled_count == 6:
            full_labeled += 1
            if normalize_text(row.get("point_all_within_15px", "")).upper() == "TRUE":
                success15 += 1
            if normalize_text(row.get("point_all_within_20px", "")).upper() == "TRUE":
                success20 += 1

        for i in range(6):
            err = parse_float(row.get(f"p{i + 1}_error_px", ""))
            if err is not None:
                point_errors[i].append(err)

    all_point_errors = [e for errors in point_errors for e in errors]
    return {
        "image_labeled_count": len(image_errors),
        "full_point_labeled_count": full_labeled,
        "image_mean_error_avg": statistics.mean(image_errors) if image_errors else None,
        "image_mean_error_median": statistics.median(image_errors) if image_errors else None,
        "image_mean_error_max": max(image_errors) if image_errors else None,
        "image_success_15px": (success15 / full_labeled) if full_labeled else None,
        "image_success_20px": (success20 / full_labeled) if full_labeled else None,
        "point_labeled_count": len(all_point_errors),
        "point_error_avg": statistics.mean(all_point_errors) if all_point_errors else None,
        "point_error_median": statistics.median(all_point_errors) if all_point_errors else None,
        "point_success_15px": (sum(1 for e in all_point_errors if e <= 15.0) / len(all_point_errors)) if all_point_errors else None,
        "point_success_20px": (sum(1 for e in all_point_errors if e <= 20.0) / len(all_point_errors)) if all_point_errors else None,
        "per_point_avg": [statistics.mean(errors) if errors else None for errors in point_errors],
    }


def fmt(value, digits=4):
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.{digits}f}"
    return str(value)


def write_csv_summary(path, direction, side, bad, points):
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["section", "metric", "value"])
        for name, summary in [("direction", direction), ("shoe_side", side), ("bad", bad)]:
            writer.writerow([name, "N", summary["n"]])
            writer.writerow([name, "correct", summary["k"]])
            writer.writerow([name, "accuracy", fmt(summary["accuracy"])])
            writer.writerow([name, "binomial_p_value_accuracy_gt_0.5", fmt(summary["p_value_gt_0_5"])])

        for key, value in points.items():
            if key == "per_point_avg":
                continue
            writer.writerow(["points", key, fmt(value)])
        for i, value in enumerate(points["per_point_avg"], start=1):
            writer.writerow(["points", f"p{i}_avg_error_px", fmt(value)])


def write_markdown(path, direction, side, bad, points):
    lines = [
        "# Experiment Result Analysis",
        "",
        "## Direction Hypothesis Test",
        "",
        f"- N: {direction['n']}",
        f"- Correct: {direction['k']}",
        f"- Accuracy: {fmt(direction['accuracy'])}",
        f"- One-sided binomial p-value, H1 accuracy > 0.5: {fmt(direction['p_value_gt_0_5'])}",
        "",
        "## Classification Summary",
        "",
        f"- Shoe side accuracy: {fmt(side['accuracy'])} ({side['k']}/{side['n']})",
        f"- BAD accuracy: {fmt(bad['accuracy'])} ({bad['k']}/{bad['n']})",
        "",
        "## Point Error Summary",
        "",
        f"- Images with at least one manual point: {points['image_labeled_count']}",
        f"- Images with all 6 manual points: {points['full_point_labeled_count']}",
        f"- Mean image point error: {fmt(points['image_mean_error_avg'])} px",
        f"- Median image point error: {fmt(points['image_mean_error_median'])} px",
        f"- Max image point error: {fmt(points['image_mean_error_max'])} px",
        f"- Image success@15px: {fmt(points['image_success_15px'])}",
        f"- Image success@20px: {fmt(points['image_success_20px'])}",
        f"- Point-level success@15px, reference only: {fmt(points['point_success_15px'])}",
        f"- Point-level success@20px, reference only: {fmt(points['point_success_20px'])}",
        "",
        "Note: point-level values are reference indicators only. Six points from the same image are not independent samples.",
    ]
    with open(path, "w", encoding="utf-8-sig") as f:
        f.write("\n".join(lines) + "\n")


def write_xlsx_summary(path, direction, side, bad, points):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "summary"
    ws.append(["section", "metric", "value"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    for name, summary in [("direction", direction), ("shoe_side", side), ("bad", bad)]:
        ws.append([name, "N", summary["n"]])
        ws.append([name, "correct", summary["k"]])
        ws.append([name, "accuracy", summary["accuracy"]])
        ws.append([name, "binomial_p_value_accuracy_gt_0.5", summary["p_value_gt_0_5"]])

    for key, value in points.items():
        if key == "per_point_avg":
            continue
        ws.append(["points", key, value])
    for i, value in enumerate(points["per_point_avg"], start=1):
        ws.append(["points", f"p{i}_avg_error_px", value])

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 48)

    wb.save(path)
    return True


def xml_escape(value):
    return (
        str(value if value is not None else "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def write_xml_excel_summary(path, direction, side, bad, points):
    rows = [["section", "metric", "value"]]
    for name, summary in [("direction", direction), ("shoe_side", side), ("bad", bad)]:
        rows.append([name, "N", summary["n"]])
        rows.append([name, "correct", summary["k"]])
        rows.append([name, "accuracy", fmt(summary["accuracy"])])
        rows.append([name, "binomial_p_value_accuracy_gt_0.5", fmt(summary["p_value_gt_0_5"])])

    for key, value in points.items():
        if key == "per_point_avg":
            continue
        rows.append(["points", key, fmt(value)])
    for i, value in enumerate(points["per_point_avg"], start=1):
        rows.append(["points", f"p{i}_avg_error_px", fmt(value)])

    parts = [
        '<?xml version="1.0"?>',
        '<?mso-application progid="Excel.Sheet"?>',
        '<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet" '
        'xmlns:o="urn:schemas-microsoft-com:office:office" '
        'xmlns:x="urn:schemas-microsoft-com:office:excel" '
        'xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">',
        '<Styles><Style ss:ID="Header"><Font ss:Bold="1"/>'
        '<Interior ss:Color="#D9EAF7" ss:Pattern="Solid"/></Style></Styles>',
        '<Worksheet ss:Name="summary"><Table>',
    ]
    for row_index, row in enumerate(rows):
        style = ' ss:StyleID="Header"' if row_index == 0 else ""
        parts.append(f"<Row{style}>")
        for value in row:
            parts.append(f'<Cell><Data ss:Type="String">{xml_escape(value)}</Data></Cell>')
        parts.append("</Row>")
    parts.extend(
        [
            "</Table>",
            '<WorksheetOptions xmlns="urn:schemas-microsoft-com:office:excel">'
            "<FreezePanes/><FrozenNoSplit/><SplitHorizontal>1</SplitHorizontal>"
            "<TopRowBottomPane>1</TopRowBottomPane></WorksheetOptions>",
            "</Worksheet></Workbook>",
        ]
    )

    with open(path, "w", encoding="utf-8-sig") as f:
        f.write("\n".join(parts))


def main():
    parser = argparse.ArgumentParser(description="Analyze labeled shoe mask experiment results.")
    parser.add_argument(
        "--input",
        default=r"C:\Users\Hyunzang\OneDrive\바탕 화면\기말텀프로젝트\마스크이미지\output\experiment_results.csv",
        help="Path to experiment_results.csv, .xls XML spreadsheet, or .xlsx file.",
    )
    parser.add_argument(
        "--output",
        default=r"C:\Users\Hyunzang\OneDrive\바탕 화면\기말텀프로젝트\마스크이미지\output\hypothesis_test_results.xlsx",
        help="Output .xlsx summary path. CSV and Markdown summaries are also written next to it.",
    )
    args = parser.parse_args()

    rows = read_table(args.input)
    if not rows:
        print("No rows found in input file.", file=sys.stderr)
        return 1

    direction = summarize_binary(rows, "pred_front_sign", "true_front_sign")
    side = summarize_binary(rows, "pred_shoe_side", "true_shoe_side")
    bad = summarize_binary(rows, "is_bad", "true_bad", normalize_bad)
    points = summarize_points(rows)

    output_dir = os.path.dirname(args.output) or "."
    os.makedirs(output_dir, exist_ok=True)
    stem, _ = os.path.splitext(args.output)
    csv_summary = stem + "_summary.csv"
    md_summary = stem + "_summary.md"
    xml_excel_summary = stem + "_summary.xls"

    wrote_xlsx = write_xlsx_summary(args.output, direction, side, bad, points)
    write_xml_excel_summary(xml_excel_summary, direction, side, bad, points)
    write_csv_summary(csv_summary, direction, side, bad, points)
    write_markdown(md_summary, direction, side, bad, points)

    print("Analysis complete")
    print(f"input: {args.input}")
    if wrote_xlsx:
        print(f"xlsx: {args.output}")
    else:
        print("xlsx: skipped because openpyxl is not installed")
    print(f"excel-compatible xls: {xml_excel_summary}")
    print(f"csv: {csv_summary}")
    print(f"markdown: {md_summary}")
    print(f"direction accuracy: {fmt(direction['accuracy'])} ({direction['k']}/{direction['n']}), p={fmt(direction['p_value_gt_0_5'])}")
    print(f"point mean error: {fmt(points['image_mean_error_avg'])} px")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
